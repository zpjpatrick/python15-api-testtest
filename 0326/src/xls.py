import json

from openpyxl import load_workbook

# 共享变量。 1： global, 类，

# class DoExcel:
def get_data():
    """获取excel的登录信息，保存到列表里面"""
    wb = load_workbook("data.xlsx")

    sheet = wb.worksheets[0]

    login_datas = []
    for row in range(2, sheet.max_row + 1):
        user = {
            "url": sheet.cell(row, 1).value,
            "method": sheet.cell(row, 2).value,
            "data": json.loads(sheet.cell(row, 3).value),
            "exp_data": sheet.cell(row, 4).value,
            "case_id" : sheet.cell(row, 7).value,
        }
        login_datas.append(user)
    wb.close()

    return login_datas


def write_result(row, result):
    wb = load_workbook("data.xlsx")

    sheet = wb.worksheets[0]

    sheet.cell(row, 6, result)

    wb.save("data.xlsx")
    wb.close()


